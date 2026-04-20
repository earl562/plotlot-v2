"""Excel renderer — assembles pro forma data into a multi-sheet .xlsx workbook.

Uses openpyxl to build a professional spreadsheet with:
- Summary sheet (property, zoning, development potential)
- Development Costs sheet (land, hard, soft, contingency)
- Revenue Projections sheet (rental income, sale proceeds)
- Financing sheet (LTV, loan, debt service, equity)
- Returns sheet (cap rate, cash-on-cash, ROI)
- Comparable Sales sheet (from comp analysis data)
- Sensitivity Matrix sheet (returns at varied assumptions)
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from plotlot.clauses.schema import AssemblyConfig, DealContext, GeneratedDocument, RenderedClause

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TOTAL_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
_LABEL_FONT = Font(name="Calibri", size=11)
_VALUE_FONT = Font(name="Calibri", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E7E5E4"),
    right=Side(style="thin", color="E7E5E4"),
    top=Side(style="thin", color="E7E5E4"),
    bottom=Side(style="thin", color="E7E5E4"),
)


def _fmt_currency(value: float) -> str:
    return f"${value:,.0f}"


def _fmt_pct(value: float) -> str:
    return f"{value:.1f}%"


def _style_header_row(ws, row: int, cols: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _style_data_row(ws, row: int, cols: int, is_total: bool = False) -> None:
    """Apply data row styling."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = _THIN_BORDER
        if is_total:
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
        else:
            cell.font = _LABEL_FONT if col == 1 else _VALUE_FONT


def _add_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list[str]],
    total_row: int | None = None,
) -> int:
    """Write a header+data table and return the next available row."""
    # Headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header_row(ws, start_row, len(headers))

    # Data rows
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        is_total = total_row is not None and (row_idx - start_row - 1) == total_row
        _style_data_row(ws, row_idx, len(headers), is_total=is_total)

    return start_row + len(rows) + 2  # next row with gap


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_summary_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 1: Property & development summary."""
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    ws.cell(row=1, column=1, value="Development Pro Forma").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=ctx.property_address or ctx.formatted_address or "—")

    row = 4
    row = _add_table(
        ws,
        row,
        ["Property Detail", "Value"],
        [
            ["Address", ctx.property_address or "—"],
            ["Municipality", ctx.municipality or "—"],
            ["County", ctx.county or "—"],
            ["State", ctx.state_code],
            ["APN / Folio", ctx.apn or "—"],
            ["Lot Size (sqft)", f"{ctx.lot_size_sqft:,.0f}" if ctx.lot_size_sqft else "—"],
            ["Year Built", str(ctx.year_built) if ctx.year_built else "—"],
        ],
    )

    row = _add_table(
        ws,
        row,
        ["Zoning / Development", "Value"],
        [
            ["Zoning District", ctx.zoning_district or "—"],
            ["Description", ctx.zoning_description or "—"],
            ["Max Height", ctx.max_height or "—"],
            ["Max Density", ctx.max_density or "—"],
            ["Max Units", str(ctx.max_units) if ctx.max_units else "—"],
            ["Governing Constraint", ctx.governing_constraint or "—"],
        ],
    )


def _build_costs_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 2: Development costs breakdown."""
    ws = wb.create_sheet("Development Costs")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    land = ctx.purchase_price or ctx.estimated_land_value or 0
    hard = ctx.hard_costs
    soft = ctx.soft_costs
    contingency = hard * 0.10 if hard else 0
    total = land + hard + soft + contingency

    rows = [
        ["Land Acquisition", _fmt_currency(land)],
        ["Hard Costs (Construction)", _fmt_currency(hard)],
        ["Soft Costs", _fmt_currency(soft)],
        ["Contingency (10%)", _fmt_currency(contingency)],
        ["Total Development Cost", _fmt_currency(total)],
    ]

    _add_table(ws, 1, ["Cost Category", "Amount"], rows, total_row=len(rows) - 1)

    # Add cost per door if available
    if ctx.cost_per_door:
        ws.cell(row=len(rows) + 3, column=1, value="Cost per Door")
        ws.cell(row=len(rows) + 3, column=2, value=_fmt_currency(ctx.cost_per_door))


def _build_revenue_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 3: Revenue projections."""
    ws = wb.create_sheet("Revenue Projections")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    gdv = ctx.gross_development_value
    margin = ctx.builder_margin
    max_land = ctx.max_land_price
    adv = ctx.adv_per_unit

    rows = [
        ["Gross Development Value (GDV)", _fmt_currency(gdv)],
        ["Builder Margin", _fmt_pct(margin * 100) if margin < 1 else _fmt_pct(margin)],
        ["Max Land Price", _fmt_currency(max_land)],
        ["After-Development Value / Unit", _fmt_currency(adv)],
    ]

    _add_table(ws, 1, ["Revenue Metric", "Value"], rows)


def _build_financing_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 4: Financing structure."""
    ws = wb.create_sheet("Financing")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    rows: list[list[str]] = [
        ["Deal Type", ctx.deal_type.value.replace("_", " ").title()],
        ["Financing Type", ctx.financing_type or "—"],
        ["Purchase Price", _fmt_currency(ctx.purchase_price)],
        ["Earnest Money", _fmt_currency(ctx.earnest_money)],
        ["Down Payment", _fmt_currency(ctx.down_payment)],
        ["Cash at Closing", _fmt_currency(ctx.cash_at_closing)],
    ]

    if ctx.existing_mortgage_balance_1:
        rows.extend(
            [
                ["Existing Mortgage Balance", _fmt_currency(ctx.existing_mortgage_balance_1)],
                ["Existing Monthly Payment", _fmt_currency(ctx.existing_mortgage_payment)],
                ["Existing Rate", _fmt_pct(ctx.existing_mortgage_rate)],
            ]
        )

    if ctx.seller_carryback_amount:
        rows.extend(
            [
                ["Seller Carryback Amount", _fmt_currency(ctx.seller_carryback_amount)],
                ["Seller Carryback Rate", _fmt_pct(ctx.seller_carryback_rate)],
                ["Seller Carryback Term", f"{ctx.seller_carryback_term_months} months"],
                ["Seller Monthly Payment", _fmt_currency(ctx.seller_carryback_payment)],
            ]
        )

    _add_table(ws, 1, ["Financing Detail", "Value"], rows)


def _build_returns_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 5: Returns analysis."""
    ws = wb.create_sheet("Returns")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    gdv = ctx.gross_development_value
    total_cost = (
        ctx.hard_costs + ctx.soft_costs + (ctx.purchase_price or ctx.estimated_land_value or 0)
    )
    profit = gdv - total_cost if gdv and total_cost else 0
    roi = (profit / total_cost * 100) if total_cost > 0 else 0

    rows = [
        ["Gross Development Value", _fmt_currency(gdv)],
        ["Total Development Cost", _fmt_currency(total_cost)],
        ["Estimated Profit", _fmt_currency(profit)],
        ["ROI", _fmt_pct(roi)],
    ]

    if ctx.max_units and ctx.max_units > 0:
        rows.append(["Profit per Unit", _fmt_currency(profit / ctx.max_units)])

    _add_table(ws, 1, ["Return Metric", "Value"], rows)


def _build_comps_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 6: Comparable sales data."""
    ws = wb.create_sheet("Comparable Sales")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    rows = [
        ["Median Price per Acre", _fmt_currency(ctx.median_price_per_acre)],
        ["Estimated Land Value", _fmt_currency(ctx.estimated_land_value)],
        ["Comparable Sales Count", str(ctx.comp_count)],
        ["Confidence Score", _fmt_pct(ctx.comp_confidence * 100) if ctx.comp_confidence else "—"],
    ]

    _add_table(ws, 1, ["Comp Metric", "Value"], rows)


def _build_sensitivity_sheet(wb: Workbook, ctx: DealContext) -> None:
    """Sheet 7: Sensitivity matrix — returns at varied assumptions.

    Rows = construction cost per sqft variations
    Cols = sale price per unit variations
    """
    ws = wb.create_sheet("Sensitivity")
    ws.column_dimensions["A"].width = 22

    if not ctx.max_units or ctx.max_units <= 0:
        ws.cell(row=1, column=1, value="Insufficient data for sensitivity analysis")
        return

    base_gdv = ctx.gross_development_value
    base_cost = ctx.hard_costs + ctx.soft_costs
    land = ctx.purchase_price or ctx.estimated_land_value or 0

    if base_gdv <= 0 or base_cost <= 0:
        ws.cell(row=1, column=1, value="Insufficient data for sensitivity analysis")
        return

    # Vary GDV by -20% to +20% in 10% steps
    gdv_factors = [0.80, 0.90, 1.00, 1.10, 1.20]
    # Vary costs by -20% to +20% in 10% steps
    cost_factors = [0.80, 0.90, 1.00, 1.10, 1.20]

    ws.cell(row=1, column=1, value="ROI Sensitivity Matrix").font = _TITLE_FONT
    ws.cell(row=2, column=1, value="Rows = Cost Factor | Cols = GDV Factor")

    # Header row (GDV factors)
    start_row = 4
    ws.cell(row=start_row, column=1, value="Cost \\ GDV")
    for col_idx, gf in enumerate(gdv_factors, 2):
        ws.cell(row=start_row, column=col_idx, value=f"GDV {gf:.0%}")
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    _style_header_row(ws, start_row, len(gdv_factors) + 1)

    # Data rows
    for row_idx, cf in enumerate(cost_factors, start_row + 1):
        ws.cell(row=row_idx, column=1, value=f"Cost {cf:.0%}")
        for col_idx, gf in enumerate(gdv_factors, 2):
            varied_gdv = base_gdv * gf
            varied_cost = base_cost * cf + land
            profit = varied_gdv - varied_cost
            roi = (profit / varied_cost * 100) if varied_cost > 0 else 0
            ws.cell(row=row_idx, column=col_idx, value=_fmt_pct(roi))
        _style_data_row(ws, row_idx, len(gdv_factors) + 1, is_total=(cf == 1.00))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def render_xlsx(
    clauses: list[RenderedClause],
    config: AssemblyConfig,
    context: DealContext,
) -> GeneratedDocument:
    """Render a pro forma spreadsheet as a multi-sheet .xlsx workbook.

    Args:
        clauses: Rendered clauses (used for metadata, not primary content).
        config: Assembly configuration.
        context: DealContext with all financial and property data.

    Returns:
        GeneratedDocument with .xlsx bytes.
    """
    wb = Workbook()

    _build_summary_sheet(wb, context)
    _build_costs_sheet(wb, context)
    _build_revenue_sheet(wb, context)
    _build_financing_sheet(wb, context)
    _build_returns_sheet(wb, context)
    _build_comps_sheet(wb, context)
    _build_sensitivity_sheet(wb, context)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)

    # Build filename
    address_part = (
        (context.property_address or context.formatted_address or "property")
        .split(",")[0]
        .replace(" ", "_")[:30]
    )
    now = datetime.now(timezone.utc)
    filename = f"PROFORMA_{address_part}_{now.strftime('%Y%m%d')}.xlsx"

    logger.info("Generated pro forma spreadsheet: %s (%d bytes)", filename, len(buf.getvalue()))
    return GeneratedDocument(
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=buf.getvalue(),
    )
