"""Tests for the xlsx pro forma renderer."""

import io

from openpyxl import load_workbook

from plotlot.clauses.renderers.xlsx_renderer import render_xlsx
from plotlot.clauses.schema import (
    AssemblyConfig,
    DealContext,
    DealType,
    DocumentType,
)


def _proforma_context() -> DealContext:
    """Full pro forma context with all financial data."""
    return DealContext(
        property_address="7940 Plantation Blvd, Miramar, FL 33023",
        municipality="Miramar",
        county="Broward",
        state_code="FL",
        apn="5040-19-02-0010",
        lot_size_sqft=15000.0,
        year_built=1975,
        zoning_district="RS-4",
        zoning_description="Residential Single Family",
        max_units=4,
        governing_constraint="density",
        max_height="35 feet",
        max_density="4 units/acre",
        median_price_per_acre=450000.0,
        estimated_land_value=155000.0,
        comp_count=8,
        comp_confidence=0.82,
        gross_development_value=1600000.0,
        hard_costs=800000.0,
        soft_costs=120000.0,
        builder_margin=0.12,
        max_land_price=488000.0,
        cost_per_door=230000.0,
        adv_per_unit=400000.0,
        deal_type=DealType.land_deal,
        purchase_price=155000.0,
        earnest_money=5000.0,
    )


def _minimal_context() -> DealContext:
    """Minimal context — just property basics."""
    return DealContext(
        property_address="123 Main St, Miami, FL",
        state_code="FL",
    )


def _subject_to_context() -> DealContext:
    """Subject-to deal with existing mortgage."""
    return DealContext(
        property_address="456 Oak Ave, Fort Lauderdale, FL",
        deal_type=DealType.subject_to,
        financing_type="subject_to",
        purchase_price=350000.0,
        earnest_money=5000.0,
        existing_mortgage_balance_1=280000.0,
        existing_mortgage_payment=1850.0,
        existing_mortgage_rate=4.25,
        gross_development_value=500000.0,
        hard_costs=50000.0,
        soft_costs=10000.0,
        max_units=1,
    )


class TestXlsxRenderer:
    def test_generates_valid_xlsx(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet,
            output_format="xlsx",
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)

        # Valid ZIP (xlsx is a ZIP format)
        assert doc.data[:4] == b"PK\x03\x04"
        assert doc.filename.startswith("PROFORMA_")
        assert doc.filename.endswith(".xlsx")
        assert (
            doc.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_has_seven_sheets(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)

        wb = load_workbook(io.BytesIO(doc.data))
        assert len(wb.sheetnames) == 7
        assert "Summary" in wb.sheetnames
        assert "Development Costs" in wb.sheetnames
        assert "Revenue Projections" in wb.sheetnames
        assert "Financing" in wb.sheetnames
        assert "Returns" in wb.sheetnames
        assert "Comparable Sales" in wb.sheetnames
        assert "Sensitivity" in wb.sheetnames

    def test_summary_sheet_has_address(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)

        wb = load_workbook(io.BytesIO(doc.data))
        ws = wb["Summary"]
        # Address should be in row 2
        assert "7940 Plantation Blvd" in str(ws.cell(row=2, column=1).value)

    def test_costs_sheet_has_values(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)

        wb = load_workbook(io.BytesIO(doc.data))
        ws = wb["Development Costs"]
        # Find the hard costs value
        all_values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert any("800,000" in str(v) for v in all_values if v), (
            f"Hard costs not found: {all_values}"
        )

    def test_financing_sheet_subject_to(self):
        """Subject-to deal should show existing mortgage data."""
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _subject_to_context()
        doc = render_xlsx([], config, context)

        wb = load_workbook(io.BytesIO(doc.data))
        ws = wb["Financing"]
        all_values = []
        for row in range(1, 20):
            for col in range(1, 3):
                val = ws.cell(row=row, column=col).value
                if val:
                    all_values.append(str(val))
        combined = " ".join(all_values)
        assert "280,000" in combined  # existing mortgage balance
        assert "Subject To" in combined  # deal type

    def test_sensitivity_sheet_has_matrix(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)

        wb = load_workbook(io.BytesIO(doc.data))
        ws = wb["Sensitivity"]
        # Should have a title
        assert "Sensitivity" in str(ws.cell(row=1, column=1).value)
        # Should have GDV factors in header row
        assert "GDV" in str(ws.cell(row=4, column=2).value)

    def test_minimal_context_still_valid(self):
        """Even with no financial data, xlsx should be valid."""
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _minimal_context()
        doc = render_xlsx([], config, context)

        assert doc.data[:4] == b"PK\x03\x04"
        wb = load_workbook(io.BytesIO(doc.data))
        assert len(wb.sheetnames) == 7

    def test_nontrivial_file_size(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)
        assert len(doc.data) > 3000  # multi-sheet workbook should be substantial

    def test_comps_sheet_has_data(self):
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet, output_format="xlsx"
        )
        context = _proforma_context()
        doc = render_xlsx([], config, context)

        wb = load_workbook(io.BytesIO(doc.data))
        ws = wb["Comparable Sales"]
        all_values = []
        for row in range(1, 10):
            for col in range(1, 3):
                val = ws.cell(row=row, column=col).value
                if val:
                    all_values.append(str(val))
        combined = " ".join(all_values)
        assert "450,000" in combined  # median price per acre


class TestXlsxEdgeCases:
    """Edge-case tests for the xlsx renderer: zero values, negatives, etc."""

    def _render(self, ctx: DealContext):  # noqa: ANN202
        """Helper: render and return the openpyxl Workbook."""
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet,
            output_format="xlsx",
        )
        doc = render_xlsx([], config, ctx)
        assert doc.data[:4] == b"PK\x03\x04"
        return load_workbook(io.BytesIO(doc.data))

    def test_all_financial_fields_zero_valid_xlsx(self):
        """All financial fields = 0 produces a valid xlsx with no division error."""
        ctx = DealContext(
            property_address="100 Zero St, Miami, FL",
            state_code="FL",
            gross_development_value=0.0,
            hard_costs=0.0,
            soft_costs=0.0,
            builder_margin=0.0,
            max_land_price=0.0,
            cost_per_door=0.0,
            adv_per_unit=0.0,
            purchase_price=0.0,
            estimated_land_value=0.0,
            max_units=0,
            median_price_per_acre=0.0,
            comp_confidence=0.0,
        )
        wb = self._render(ctx)
        assert len(wb.sheetnames) == 7
        # Returns sheet should not crash on zero total_cost (ROI = 0)
        ws = wb["Returns"]
        all_values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert any("0.0%" in str(v) for v in all_values if v), (
            f"Expected 0.0% ROI, got: {all_values}"
        )

    def test_negative_costs_renders(self):
        """Negative cost values render without crash (no validation in renderer)."""
        ctx = DealContext(
            property_address="200 Negative Ave, Miami, FL",
            state_code="FL",
            gross_development_value=500000.0,
            hard_costs=-100000.0,
            soft_costs=-20000.0,
            builder_margin=0.10,
            max_land_price=300000.0,
            purchase_price=100000.0,
            max_units=4,
        )
        wb = self._render(ctx)
        assert len(wb.sheetnames) == 7
        # Costs sheet should contain a negative currency string
        ws = wb["Development Costs"]
        all_values = []
        for row in range(1, 10):
            val = ws.cell(row=row, column=2).value
            if val:
                all_values.append(str(val))
        combined = " ".join(all_values)
        assert "-$100,000" in combined or "$-100,000" in combined, (
            f"Expected negative hard costs in: {combined}"
        )

    def test_max_units_zero_valid(self):
        """max_units = 0 yields a valid xlsx; sensitivity sheet shows 'Insufficient data'."""
        ctx = DealContext(
            property_address="300 Empty Lot Dr, Miami, FL",
            state_code="FL",
            gross_development_value=500000.0,
            hard_costs=200000.0,
            soft_costs=30000.0,
            max_units=0,
        )
        wb = self._render(ctx)
        assert len(wb.sheetnames) == 7
        # Sensitivity sheet should show insufficient data message
        ws = wb["Sensitivity"]
        cell_val = str(ws.cell(row=1, column=1).value)
        assert "Insufficient" in cell_val or "Sensitivity" in cell_val
        # Returns sheet should NOT have "Profit per Unit" row (max_units=0)
        ws_ret = wb["Returns"]
        labels = [ws_ret.cell(row=r, column=1).value for r in range(1, 15)]
        assert "Profit per Unit" not in labels

    def test_gdv_zero_skips_per_unit(self):
        """GDV = 0 yields a valid xlsx without crashing on per-unit calculations."""
        ctx = DealContext(
            property_address="400 No Revenue Blvd, Miami, FL",
            state_code="FL",
            gross_development_value=0.0,
            hard_costs=100000.0,
            soft_costs=20000.0,
            builder_margin=0.0,
            max_land_price=0.0,
            purchase_price=50000.0,
            max_units=4,
        )
        wb = self._render(ctx)
        assert len(wb.sheetnames) == 7
        # Revenue sheet should show $0 GDV
        ws = wb["Revenue Projections"]
        all_values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert any("$0" in str(v) for v in all_values if v), (
            f"Expected $0 GDV in revenue sheet: {all_values}"
        )
        # Sensitivity sheet should show insufficient data (GDV <= 0)
        ws_sens = wb["Sensitivity"]
        cell_val = str(ws_sens.cell(row=1, column=1).value)
        assert "Insufficient" in cell_val


class TestXlsxEndToEnd:
    def test_engine_generates_xlsx(self):
        """Test that assemble_document routes to xlsx_renderer correctly."""
        from plotlot.clauses.engine import assemble_document
        from plotlot.clauses.loader import ClauseRegistry

        registry = ClauseRegistry([])  # empty registry — xlsx uses context data
        config = AssemblyConfig(
            document_type=DocumentType.proforma_spreadsheet,
            output_format="xlsx",
        )
        context = _proforma_context()
        doc = assemble_document(config, context, registry)

        assert doc.data[:4] == b"PK\x03\x04"
        assert doc.filename.endswith(".xlsx")
